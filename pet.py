import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import List

# ==== Настройки ====
COUNTRY_CODES = {
    "Грузия": "GE",
    "Беларусь": "BY",
    "Азербайджан": "AZ",
    "Кыргызстан": "KG",
    "Армения": "AM",
    "Казахстан": "KZ",
    "Таджикистан": "TJ",
    "Узбекистан": "UZ",
}

# MPID для стран
COUNTRY_MPID = {
    "Армения": 25,
    "Беларусь": 3,
    "Казахстан": 2,
    "Кыргызстан": 23,
    "Узбекистан": 30,
    "Грузия": 1004,
    "Азербайджан": 1003,
}

API_URL = "https://calendarific.com/api/v2/holidays"
API_KEY = "ETEJbeuaHCz8C7WH4xSAJ408oMfug05S"  # <- ключ


def fetch_public_holidays(country_code: str, year: int) -> set:
    params = {
        "api_key": API_KEY,
        "country": country_code,
        "year": year,
        "type": "national",
    }
    resp = requests.get(API_URL, params=params, timeout=30)
    resp.raise_for_status()
    data = resp.json()
    holidays = set()
    for item in data.get("response", {}).get("holidays", []):
        iso = item.get("date", {}).get("iso")
        if iso:
            holidays.add(iso[:10])
    return holidays


def build_table(year: int, selected_countries: list[str]) -> pd.DataFrame:
    date_range = pd.date_range(f"{year}-01-01", f"{year}-12-31", freq="D")
    date_cols = [d.strftime("%d.%m.%Y") for d in date_range]

    rows = []
    for country_name, code in COUNTRY_CODES.items():
        if country_name not in selected_countries:
            continue
        holidays_iso = fetch_public_holidays(code, year)

        row = {
            "Country": country_name,                       # <<< ИЗМЕНЕНО (теперь первый)
            "MPID": COUNTRY_MPID.get(country_name, None),  # <<< ИЗМЕНЕНО (теперь второй)
        }
        for d in date_range:
            iso = d.strftime("%Y-%m-%d")
            if iso in holidays_iso:
                val = "holiday"
            elif d.weekday() >= 5:
                val = "day off"
            else:
                val = "work day"
            row[d.strftime("%d.%m.%Y")] = val

        rows.append(row)

    # порядок колонок: Country → MPID → даты
    df = pd.DataFrame(rows, columns=["Country", "MPID"] + date_cols)  # <<< ИЗМЕНЕНО
    return df


def apply_colors(outfile: str):
    wb = load_workbook(outfile)
    ws = wb.active

    fills = {
        "holiday": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Зеленый
        "day off": PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"), # Серый
        "work day": PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid"),# Красный
    }

    # сдвигаем, т.к. теперь первые два столбца — Country и MPID
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=ws.max_column):  # <<< ИЗМЕНЕНО
        for cell in row:
            val = cell.value
            if val in fills:
                cell.fill = fills[val]

    wb.save(outfile)


def main():
    year = 2026
    df = build_table(year, list(COUNTRY_MPID.keys()))
    outfile = f"work_calendar_{year}.xlsx"
    df.to_excel(outfile, index=False)
    apply_colors(outfile)
    print(f"Готово! Файл сохранен: {outfile}")


if __name__ == "__main__":
    main()